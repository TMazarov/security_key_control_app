import PySimpleGUI as sg
import openpyxl

# Set theme for the window
sg.theme("Topanga")

# Excel file path
EXCEL_FILE = "saflok_security_key_control.xlsx"

# load the existing Excel file
book = openpyxl.load_workbook(EXCEL_FILE)
sheet = book.active

layout = [
    [sg.Text("Date:", size=(16, 1)), sg.InputText(key="Date:")],
    [sg.Text("User:", size=(16, 1)), sg.InputText(key="User:")],
    [sg.Text("Section:", size=(16, 1)), sg.InputText(key="Section:")],
    [sg.Text("ID No.", size=(16, 1)), sg.InputText(key="ID No.")],
    [sg.Text("Reason for Cancelling", size=(16, 1)), sg.InputText(key="Reason for Cancelling")],
    [sg.Text("Duty Security", size=(16, 1)), sg.InputText(key="Duty Security")],
    [sg.Submit(), sg.Button("Clear"), sg.Exit()]
]

window = sg.Window("GHS Saflok Security Key Control", layout)


def clear_input():
    for key in values:
        window[key]("")
    return None


while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == "Exit":
        break

    if event == "Clear":
        clear_input()

    if event == "Submit":

        # find the next available row in the sheet
        next_row = sheet.max_row + 1

        # write the data on the sheet
        sheet.cell(row=next_row, column=1).value = values["Date:"]
        sheet.cell(row=next_row, column=2).value = values["User:"]
        sheet.cell(row=next_row, column=3).value = values["Section:"]
        sheet.cell(row=next_row, column=4).value = values["ID No."]
        sheet.cell(row=next_row, column=5).value = values["Reason for Cancelling"]
        sheet.cell(row=next_row, column=6).value = values["Duty Security"]

        # Save changes
        book.save(EXCEL_FILE)

        sg.popup("Information Saved on Excel Sheet!")

        clear_input()

window.close()
